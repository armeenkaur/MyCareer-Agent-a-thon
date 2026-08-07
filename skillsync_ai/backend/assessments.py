from __future__ import annotations

from typing import Any
from ..agents.llm import normalize_proficiency
from ..core.config import PROFICIENCY_ORDER, PROFICIENCY_VALUE, UPLOAD_DIR
from ..core.logging_setup import get_logger
from ..data_sources import CAREER_MOVE_LABELS
from ..database import Database, FEEDBACK_QUESTION, KUDOS_PRESET, PHASES, PHASE_FREE_ROLES, ist_today, utc_now
from .errors import BackendError
log = get_logger('skillsync.backend')

_CAREER_MOVE_HEADERS = {
    "career move",
    "career_move",
    "career recommendation",
    "career_recommendation",
}
_LOB_NOTE_HEADERS = {
    "lob change note",
    "lob_change_note",
    "career recommendation note",
    "career_recommendation_note",
    "lob note",
}

class AssessmentsMixin:
    def assessment(self, employee_code: str, role: str) -> dict[str, Any] | None:
        with self.db.connect() as connection:
            row = connection.execute(
                "SELECT * FROM assessments WHERE employee_code=? AND assessor_role=?", (employee_code, role)
            ).fetchone()
            if not row:
                return None
            ratings = connection.execute(
                "SELECT competency,proficiency,note FROM assessment_ratings WHERE assessment_id=? ORDER BY competency",
                (row["id"],),
            ).fetchall()
        result = dict(row)
        result["ratings"] = {rating["competency"]: rating["proficiency"] for rating in ratings}
        result["notes"] = {rating["competency"]: rating["note"] for rating in ratings if rating["note"]}
        result["career_recommendation"] = str(result.get("career_recommendation") or "")
        result["career_recommendation_note"] = str(result.get("career_recommendation_note") or "")
        return result


    def career_move_options(self, user: dict[str, Any], employee_code: str) -> dict[str, Any]:
        if user["role"] not in {"zm", "rd", "admin"}:
            raise BackendError("ZM, RD, or Admin access required.", "forbidden", 403)
        if user["role"] in {"zm", "rd"}:
            self._assert_employee_scope(user, employee_code)
        employee = self.employee(employee_code)
        return {
            "employee_code": employee_code,
            "question": "What career move do you recommend for the employee?",
            "options": self.data.manager_career_move_options(employee),
        }


    def save_assessment(
        self,
        user: dict[str, Any],
        employee_code: str,
        ratings: dict[str, str],
        notes: dict[str, str] | None = None,
        submit: bool = False,
        career_recommendation: str = "",
        require_career_move: bool = True,
        career_recommendation_note: str = "",
    ) -> dict[str, Any]:
        role = user["role"]
        if role not in {"zm", "rd"}:
            raise BackendError("ZM or RD access required.", "forbidden", 403)
        if not self.phase_is_open(role):
            raise BackendError("Assessment phase is closed.", "phase_closed", 403)
        self._assert_employee_scope(user, employee_code)
        if role == "rd":
            zm_assessment = self.assessment(employee_code, "zm")
            if not zm_assessment or zm_assessment["status"] != "submitted":
                raise BackendError(
                    "ZM assessment must be submitted before RD validation.",
                    "zm_assessment_required",
                    409,
                )
        self._validate_ratings(ratings, complete=submit)
        recommendation = str(career_recommendation or "").strip().lower()
        move_note = str(career_recommendation_note or "").strip()
        if recommendation != "lob_change":
            move_note = ""
        if submit and require_career_move:
            self._validate_career_recommendation(employee_code, recommendation, move_note)
        elif recommendation:
            self._validate_career_recommendation(employee_code, recommendation, move_note)
        if submit and role == "zm":
            self._validate_ai_override_notes(employee_code, ratings, notes or {}, "zm_suggested_rating")
        # RD may deviate from AI suggestion without a required note.
        existing = self.assessment(employee_code, role)
        if existing and existing["status"] == "submitted":
            raise BackendError("Submitted assessment is locked.", "assessment_locked", 409)
        if not recommendation and existing:
            recommendation = str(existing.get("career_recommendation") or "")
            if recommendation == "lob_change" and not move_note:
                move_note = str(existing.get("career_recommendation_note") or "")
        now = utc_now()
        with self.db.transaction() as connection:
            connection.execute(
                """
                INSERT INTO assessments(
                    employee_code,assessor_role,assessor_login_id,status,created_at,updated_at,submitted_at,
                    career_recommendation,career_recommendation_note
                )
                VALUES(?,?,?,?,?,?,?,?,?)
                ON CONFLICT(employee_code,assessor_role) DO UPDATE SET
                    assessor_login_id=excluded.assessor_login_id,status=excluded.status,
                    updated_at=excluded.updated_at,submitted_at=excluded.submitted_at,
                    career_recommendation=excluded.career_recommendation,
                    career_recommendation_note=excluded.career_recommendation_note
                """,
                (
                    employee_code,
                    role,
                    user["login_id"],
                    "submitted" if submit else "draft",
                    now,
                    now,
                    now if submit else None,
                    recommendation,
                    move_note,
                ),
            )
            assessment_id = connection.execute(
                "SELECT id FROM assessments WHERE employee_code=? AND assessor_role=?", (employee_code, role)
            ).fetchone()[0]
            for competency, proficiency in ratings.items():
                connection.execute(
                    """
                    INSERT INTO assessment_ratings(assessment_id,competency,proficiency,note) VALUES(?,?,?,?)
                    ON CONFLICT(assessment_id,competency) DO UPDATE SET proficiency=excluded.proficiency,note=excluded.note
                    """,
                    (assessment_id, competency, proficiency, (notes or {}).get(competency, "")),
                )
        if submit and self.phase_progress(role)["is_complete"]:
            with self.db.transaction() as connection:
                connection.execute("UPDATE phases SET status='complete', closed_at=? WHERE phase=?", (utc_now(), role))
        if submit and role == "rd":
            try:
                self.precompute_recommendations(employee_code)
            except Exception:  # noqa: BLE001
                log.exception("Course precomputation failed employee=%s", employee_code)
        return self.assessment(employee_code, role) or {}

    def _validate_ai_override_notes(
        self,
        employee_code: str,
        ratings: dict[str, str],
        notes: dict[str, str],
        suggestion_key: str,
    ) -> None:
        """When assessor overrides an AI suggested rating, a note is required for that competency."""
        missing: list[str] = []
        for competency, proficiency in ratings.items():
            cached = self._cached_evidence(employee_code, competency)
            if not isinstance(cached, dict):
                continue
            suggested = str(cached.get(suggestion_key) or "").strip()
            if suggested not in {"Beginner", "Intermediate", "Proficient", "Advanced"}:
                continue
            if proficiency == suggested:
                continue
            note = str(notes.get(competency) or "").strip()
            if not note:
                missing.append(competency)
        if missing:
            raise BackendError(
                "Add a note for each competency where your rating differs from the AI suggestion: "
                + ", ".join(missing),
                "note_required",
                422,
            )

    def download_assessment_template(self, user: dict[str, Any]) -> tuple[bytes, str]:
        """Excel template: all scoped reportees, AI suggestions, rating dropdowns, skill rubric."""
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        role = user.get("role")
        if role not in {"zm", "rd"}:
            raise BackendError("ZM or RD access required.", "forbidden", 403)
        if not self.phase_is_open(role):
            raise BackendError("Assessment phase is closed.", "phase_closed", 403)

        rows = self._template_employee_rows(user)
        wb = Workbook()
        ws = wb.active
        ws.title = "Ratings"

        # Pair each skill: AI suggested (read-only guide) + Your rating (dropdown).
        headers = ["Employee Code", "Employee Name"]
        rating_cols: list[int] = []  # 1-based Excel column indexes for dropdowns
        for competency in self.competencies:
            headers.append(f"{competency} (AI suggested)")
            headers.append(competency)
            rating_cols.append(len(headers))  # competency rating column
        headers.append("Career Move")
        headers.append("LOB change note")
        career_col = len(headers) - 1
        lob_note_col = len(headers)
        ws.append(headers)

        header_font = Font(bold=True)
        ai_fill = PatternFill("solid", fgColor="D5E3FF")
        submitted_fill = PatternFill("solid", fgColor="F3F4F6")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        for row in rows:
            code = row["employee_code"]
            existing = self.assessment(code, role)
            ratings = (existing or {}).get("ratings") or {}
            submitted = (existing or {}).get("status") == "submitted"
            ai_map = self._template_ai_ratings(user, code, role)
            move_id = str((existing or {}).get("career_recommendation") or "").strip()
            move_label = CAREER_MOVE_LABELS.get(move_id, move_id)
            move_note = str((existing or {}).get("career_recommendation_note") or "")
            line: list[Any] = [code, row.get("name") or ""]
            for competency in self.competencies:
                ai_level = ai_map.get(competency) or "Intermediate"
                saved = ratings.get(competency) or ""
                # Prefill editable rating with AI so managers can change; keep submitted as-is.
                if submitted:
                    your_level = saved
                else:
                    your_level = saved or ai_level
                line.append(ai_level)
                line.append(your_level)
            line.append(move_label)
            line.append(move_note if move_id == "lob_change" else "")
            ws.append(line)
            excel_row = ws.max_row
            if submitted:
                for col in range(1, len(headers) + 1):
                    ws.cell(excel_row, col).fill = submitted_fill
            # Tint AI columns
            for index, _competency in enumerate(self.competencies):
                ai_col = 3 + index * 2
                ws.cell(excel_row, ai_col).fill = ai_fill

        if rows:
            first_data = 2
            last_data = 1 + len(rows)
            levels = ",".join(PROFICIENCY_ORDER)
            for col_index in rating_cols:
                col = get_column_letter(col_index)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{levels}"',
                    allow_blank=True,
                    showDropDown=False,
                )
                dv.error = "Pick Beginner, Intermediate, Proficient, or Advanced"
                dv.errorTitle = "Invalid proficiency"
                ws.add_data_validation(dv)
                dv.add(f"{col}{first_data}:{col}{last_data}")
            move_labels = ",".join(CAREER_MOVE_LABELS.values())
            move_dv = DataValidation(
                type="list",
                formula1=f'"{move_labels}"',
                allow_blank=True,
                showDropDown=False,
            )
            move_dv.error = "Pick a career move from the list"
            move_dv.errorTitle = "Invalid career move"
            ws.add_data_validation(move_dv)
            move_dv.add(f"{get_column_letter(career_col)}{first_data}:{get_column_letter(career_col)}{last_data}")

        ws.freeze_panes = "C2"
        ws.row_dimensions[1].height = 36
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 28
        for index in range(3, len(headers) + 1):
            width = 28 if index in {career_col, lob_note_col} else 18
            ws.column_dimensions[get_column_letter(index)].width = width

        # Rubric / definitions sheet
        rubric = wb.create_sheet("Skill Definitions")
        rubric.append(
            ["Competency", "Skill definition", "Beginner", "Intermediate", "Proficient", "Advanced"]
        )
        for cell in rubric[1]:
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        by_skill = {row["skill"]: row for row in self.data.competencies}
        for competency in self.competencies:
            meta = by_skill.get(competency) or {}
            levels = self.data.level_definitions.get(competency) or {}
            rubric.append(
                [
                    competency,
                    meta.get("definition") or "",
                    levels.get("Beginner") or "",
                    levels.get("Intermediate") or "",
                    levels.get("Proficient") or "",
                    levels.get("Advanced") or "",
                ]
            )
        for row in rubric.iter_rows(min_row=2, max_row=rubric.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        rubric.column_dimensions["A"].width = 28
        rubric.column_dimensions["B"].width = 40
        for letter in ("C", "D", "E", "F"):
            rubric.column_dimensions[letter].width = 36
        rubric.freeze_panes = "A2"

        moves = wb.create_sheet("Career Move Options")
        moves.append(["Option", "ID", "Notes"])
        for cell in moves[1]:
            cell.font = header_font
        for move_id, label in CAREER_MOVE_LABELS.items():
            note = ""
            if move_id in {"kam", "zm"}:
                note = "Only valid when suggested for that employee (role/grade). Upload rejects if not allowed."
            elif move_id == "lob_change":
                note = "Requires LOB change note on the Ratings sheet."
            moves.append([label, move_id, note])
        moves.column_dimensions["A"].width = 32
        moves.column_dimensions["B"].width = 14
        moves.column_dimensions["C"].width = 70

        # Instructions sheet (keep Ratings sheet clean for upload parsing)
        guide = wb.create_sheet("Instructions", 0)
        guide.append(["Bulk assessment template"])
        guide.append([])
        guide.append(["How to use"])
        guide.append(["1. Open the Ratings sheet — every reportee in your scope is listed (Name + Employee Code)."])
        guide.append(["2. Blue columns = AI suggested rating (guide). Pre-filled for every skill."])
        guide.append(["3. White competency columns = your rating (pre-filled from AI). Change via dropdown as needed."])
        guide.append(["4. If AI evidence is thin, cells start from role ideal proficiency — still editable."])
        guide.append(["5. Career Move column — pick from dropdown. Required to submit when all seven skills are filled."])
        guide.append(["6. If Career Move = LOB change, fill LOB change note (required on submit)."])
        guide.append(["7. Leave a rating blank to skip that skill. Fill all seven + career move to submit; partial rows save as draft."])
        guide.append(["8. Grey rows = already submitted (shown for reference; upload ignores them)."])
        guide.append(["9. Open Skill Definitions and Career Move Options for meanings and allowed paths."])
        if role == "rd":
            guide.append(["10. RD template only lists employees whose ZM assessment is already submitted."])
        guide.append([])
        guide.append(["Upload the Ratings sheet back via Upload Excel."])
        guide.column_dimensions["A"].width = 110

        buffer = BytesIO()
        wb.save(buffer)
        filename = f"{role.upper()}_ratings_template.xlsx"
        return buffer.getvalue(), filename


    def _template_ai_ratings(self, user: dict[str, Any], employee_code: str, role: str) -> dict[str, str]:
        """Resolve AI levels for template — never leave blank.

        Order: stored suggestion (any curator version) → live UI agents → voice AI → role ideal.
        """
        ideals = self.data.ideal_for_employee(employee_code) or {}
        out: dict[str, str] = {}

        # Fast path: reuse any stored AI rating without re-running agents.
        for competency in self.competencies:
            raw = self._evidence_json_raw(employee_code, competency) or {}
            if role == "zm":
                level = str(raw.get("zm_suggested_rating") or raw.get("suggested_rating") or "").strip()
            else:
                level = str(raw.get("suggested_rating") or "").strip()
            if level in PROFICIENCY_ORDER:
                out[competency] = level

        missing = [c for c in self.competencies if out.get(c) not in PROFICIENCY_ORDER]
        if missing:
            try:
                if role == "zm":
                    bundle = self.zm_assessment_evidence(user, employee_code)
                else:
                    bundle = self.rd_validation_context(user, employee_code)
                evidence = bundle.get("evidence") or {}
                for competency in missing:
                    row = evidence.get(competency) or {}
                    level = str(row.get("suggested_rating") or "").strip()
                    if level in PROFICIENCY_ORDER:
                        out[competency] = level
            except Exception:  # noqa: BLE001
                log.exception("AI template prefill failed employee=%s role=%s", employee_code, role)

        voice_ai = {
            row["competency"]: row.get("ai_proficiency")
            for row in self.roleplays(employee_code, include_private=True)
            if row.get("ai_proficiency")
        }
        for competency in self.competencies:
            if out.get(competency) in PROFICIENCY_ORDER:
                continue
            level = str(voice_ai.get(competency) or "").strip()
            if level not in PROFICIENCY_ORDER:
                level = str(ideals.get(competency) or "").strip()
            if level not in PROFICIENCY_ORDER:
                level = "Intermediate"
            out[competency] = level
        return out


    def upload_assessment_workbook(
        self,
        user: dict[str, Any],
        filename: str,
        payload: bytes,
    ) -> dict[str, Any]:
        """Parse uploaded ratings workbook/CSV and apply via save_assessment."""
        role = user.get("role")
        if role not in {"zm", "rd"}:
            raise BackendError("ZM or RD access required.", "forbidden", 403)
        if not self.phase_is_open(role):
            raise BackendError("Assessment phase is closed.", "phase_closed", 403)
        if not payload:
            raise BackendError("Uploaded file is empty.")

        records = self._parse_assessment_upload(filename, payload)
        scope = {row["employee_code"] for row in self.scoped_employees(user)}
        applied: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []

        for record in records:
            code = record["employee_code"]
            ratings = record["ratings"]
            career_recommendation = str(record.get("career_recommendation") or "")
            career_note = str(record.get("career_recommendation_note") or "")
            if code not in scope:
                errors.append({"employee_code": code, "message": "Outside your reporting scope."})
                continue
            if not ratings:
                skipped.append({"employee_code": code, "message": "No proficiency values filled."})
                continue
            existing = self.assessment(code, role)
            if existing and existing.get("status") == "submitted":
                skipped.append({"employee_code": code, "message": "Already submitted and locked."})
                continue
            if role == "rd":
                zm = self.assessment(code, "zm")
                if not zm or zm.get("status") != "submitted":
                    errors.append({"employee_code": code, "message": "ZM assessment not submitted yet."})
                    continue
            submit = set(ratings.keys()) == set(self.competencies)
            try:
                result = self.save_assessment(
                    user,
                    code,
                    ratings,
                    notes={},
                    submit=submit,
                    career_recommendation=career_recommendation,
                    career_recommendation_note=career_note,
                    require_career_move=submit,
                )
                applied.append(
                    {
                        "employee_code": code,
                        "status": result.get("status"),
                        "ratings_count": len(ratings),
                        "submitted": submit,
                        "career_recommendation": result.get("career_recommendation") or "",
                    }
                )
            except BackendError as exc:
                errors.append({"employee_code": code, "message": exc.message})

        return {
            "applied": applied,
            "skipped": skipped,
            "errors": errors,
            "summary": {
                "applied": len(applied),
                "skipped": len(skipped),
                "errors": len(errors),
                "total_rows": len(records),
            },
        }


    def _template_employee_rows(self, user: dict[str, Any]) -> list[dict[str, Any]]:
        """All reportees in scope (name + ID). RD still needs ZM submitted."""
        role = user["role"]
        rows = []
        for employee in sorted(self.employee_summaries(user), key=lambda row: str(row.get("name") or "").lower()):
            if role == "rd" and employee.get("zm_status") != "submitted":
                continue
            rows.append(employee)
        return rows


    def _parse_assessment_upload(self, filename: str, payload: bytes) -> list[dict[str, Any]]:
        name = str(filename or "").lower()
        if name.endswith(".csv"):
            return self._parse_assessment_csv(payload)
        return self._parse_assessment_xlsx(payload)


    def _parse_assessment_xlsx(self, payload: bytes) -> list[dict[str, Any]]:
        from io import BytesIO

        from openpyxl import load_workbook

        try:
            wb = load_workbook(BytesIO(payload), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001
            raise BackendError("Could not read Excel file. Upload a valid .xlsx template.") from exc
        # Prefer Ratings sheet when present (new templates put Instructions first).
        ws = wb["Ratings"] if "Ratings" in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise BackendError("Workbook is empty.")
        return self._records_from_header_rows(header_row, rows_iter)


    def _parse_assessment_csv(self, payload: bytes) -> list[dict[str, Any]]:
        import csv
        from io import StringIO

        try:
            text = payload.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise BackendError("CSV must be UTF-8 encoded.") from exc
        reader = csv.reader(StringIO(text))
        try:
            header_row = next(reader)
        except StopIteration:
            raise BackendError("CSV is empty.")
        return self._records_from_header_rows(header_row, reader)


    def _records_from_header_rows(self, header_row: Any, data_rows: Any) -> list[dict[str, Any]]:
        headers = [str(cell or "").strip() for cell in header_row]
        if not headers or headers[0].lower() not in {"employee code", "employee_code", "emp code", "emp_code"}:
            raise BackendError("First column must be Employee Code.")
        # Map competency rating headers; skip AI-suggested guide columns.
        comp_index: dict[int, str] = {}
        career_index: int | None = None
        lob_note_index: int | None = None
        lower_map = {name.lower(): name for name in self.competencies}
        for index, header in enumerate(headers):
            if index < 2:
                continue
            key = header.lower().strip()
            if "(ai suggested)" in key or key.endswith("[ai suggested]"):
                continue
            if key in _CAREER_MOVE_HEADERS:
                career_index = index
                continue
            if key in _LOB_NOTE_HEADERS:
                lob_note_index = index
                continue
            if key in lower_map:
                comp_index[index] = lower_map[key]
            elif header in self.competencies:
                comp_index[index] = header
        if not comp_index:
            raise BackendError("No competency columns found. Download a fresh template.")

        records: list[dict[str, Any]] = []
        for raw in data_rows:
            cells = list(raw)
            if not cells:
                continue
            code = str(cells[0] or "").strip()
            if not code or code.lower().startswith("instruction"):
                continue
            if code.lower() == "instructions":
                break
            ratings: dict[str, str] = {}
            for index, competency in comp_index.items():
                if index >= len(cells):
                    continue
                level = normalize_proficiency(cells[index])
                if level:
                    ratings[competency] = level
            career = ""
            if career_index is not None and career_index < len(cells):
                career = self._normalize_career_move_value(cells[career_index])
            lob_note = ""
            if lob_note_index is not None and lob_note_index < len(cells):
                lob_note = str(cells[lob_note_index] or "").strip()
            records.append(
                {
                    "employee_code": code,
                    "ratings": ratings,
                    "career_recommendation": career,
                    "career_recommendation_note": lob_note,
                }
            )
        if not records:
            raise BackendError("No employee rows found in the upload.")
        return records

    @staticmethod
    def _normalize_career_move_value(value: Any) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        lower = raw.lower()
        if lower in CAREER_MOVE_LABELS:
            return lower
        for move_id, label in CAREER_MOVE_LABELS.items():
            if label.lower() == lower:
                return move_id
        return lower

    def reset_manager_assessments(
        self,
        admin: dict[str, Any],
        employee_code: str,
        scope: str = "both",
    ) -> dict[str, Any]:
        """Admin unlock: clear ZM and/or RD drafts/submissions so managers can re-enter."""
        if admin.get("role") != "admin":
            raise BackendError("Admin access required.", "forbidden", 403)
        employee = self.employee(employee_code)
        wanted = str(scope or "both").strip().lower()
        if wanted not in {"zm", "rd", "both"}:
            raise BackendError("scope must be zm, rd, or both.", "invalid_scope", 422)
        # ZM wipe invalidates RD validation — always drop RD with ZM.
        roles: list[str]
        if wanted == "rd":
            roles = ["rd"]
        elif wanted == "zm":
            roles = ["zm", "rd"]
        else:
            roles = ["zm", "rd"]

        cleared: list[str] = []
        with self.db.transaction() as connection:
            for role in roles:
                row = connection.execute(
                    "SELECT id FROM assessments WHERE employee_code=? AND assessor_role=?",
                    (employee_code, role),
                ).fetchone()
                if not row:
                    continue
                connection.execute("DELETE FROM assessments WHERE id=?", (row["id"],))
                cleared.append(role)
            if "rd" in roles:
                connection.execute(
                    "DELETE FROM course_recommendations WHERE employee_code=?",
                    (employee_code,),
                )
                connection.execute(
                    "DELETE FROM other_source_recommendations WHERE employee_code=?",
                    (employee_code,),
                )
            if cleared:
                connection.execute(
                    "DELETE FROM curated_evidence WHERE employee_code=?",
                    (employee_code,),
                )

        self._audit(
            employee_code,
            "admin_assessments",
            "reset",
            f"Admin {admin.get('employee_code') or admin.get('name') or 'admin'} reset manager assessments",
            {"scope": wanted, "cleared": cleared},
            "reset",
        )
        zm = self.assessment(employee_code, "zm")
        rd = self.assessment(employee_code, "rd")
        return {
            "status": "reset",
            "employee_code": employee["employee_code"],
            "scope": wanted,
            "cleared": cleared,
            "zm_status": zm["status"] if zm else "not_started",
            "rd_status": rd["status"] if rd else "not_started",
        }

