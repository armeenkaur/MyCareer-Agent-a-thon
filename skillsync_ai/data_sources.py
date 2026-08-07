from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from .core.config import SOURCE_FILES
from .core.utils import clean, is_kam_title, role_level_key


CAREER_MOVE_ALWAYS = (
    {"id": "bdfe", "label": "BDFE"},
    {"id": "category", "label": "Category"},
    {"id": "continue", "label": "Continue in Current Profile"},
    {"id": "lob_change", "label": "LOB change"},
)
CAREER_MOVE_LABELS = {
    "kam": "KAM",
    "zm": "ZM",
    "bdfe": "BDFE",
    "category": "Category",
    "continue": "Continue in Current Profile",
    "lob_change": "LOB change",
}


class WorkbookData:
    """Loads the attached workbooks as configurable master and evidence data."""

    def __init__(self) -> None:
        self.generated_at = datetime.now()
        self.competencies = self._load_competencies()
        self.level_definitions = self._load_level_definitions()
        self.roleplay_links = self._load_roleplay_links()
        self.employees = self._load_employees()
        self.tna = self._load_tna()
        self.appraisal = self._load_appraisal()
        self.interview = self._load_interview()
        self.amber = self._load_amber()
        self.courses = self._load_courses()
        self.career_suggestions = self._load_career_suggestions()

    def manager_accounts(self) -> list[dict[str, str]]:
        accounts: dict[str, dict[str, str]] = {}
        for employee in self.employees.values():
            code = employee.get("manager_code", "")
            if code:
                accounts[code] = {"code": code, "name": employee.get("manager", "")}
        return sorted(accounts.values(), key=lambda row: row["code"])

    def rd_accounts(self) -> list[dict[str, str]]:
        accounts: dict[str, dict[str, str]] = {}
        for employee in self.employees.values():
            code = employee.get("rd_code", "")
            if code:
                accounts[code] = {"code": code, "name": employee.get("rd", "")}
        return sorted(accounts.values(), key=lambda row: row["code"])

    def ideal_for_employee(self, emp_code: str) -> dict[str, str]:
        employee = self.employees.get(emp_code, {})
        key = role_level_key(
            employee.get("designation", ""),
            employee.get("level", ""),
            employee.get("role_name") or employee.get("role") or "",
        )
        return self.ideal_for_role_key(key)

    def ideal_for_role_key(self, key: str) -> dict[str, str]:
        ideal: dict[str, str] = {}
        for row in self.competencies:
            ideal[row["skill"]] = row["ideals"].get(key) or row["ideals"].get("BDM (RL2-3)") or "Intermediate"
        return ideal

    def career_suggestion_flags(self, employee: dict[str, Any]) -> dict[str, str]:
        """Table 2 flags for KAM/ZM: yes | grey | hide."""
        role = (
            "KAM"
            if is_kam_title(
                employee.get("role_name", ""),
                employee.get("role", ""),
                employee.get("designation", ""),
            )
            else "BDM"
        )
        grade = clean(employee.get("grade") or employee.get("level") or "")
        key = f"{role} ({grade})" if grade else role
        return dict(self.career_suggestions.get(key) or {"kam": "hide", "zm": "hide"})

    def manager_career_move_options(self, employee: dict[str, Any]) -> list[dict[str, str]]:
        """Options for ZM/RD career-move question (Table 2 + always-on paths)."""
        flags = self.career_suggestion_flags(employee)
        options: list[dict[str, str]] = []
        for move_id in ("kam", "zm"):
            flag = flags.get(move_id) or "hide"
            if flag in {"yes", "grey"}:
                options.append({"id": move_id, "label": CAREER_MOVE_LABELS[move_id]})
        options.extend(dict(row) for row in CAREER_MOVE_ALWAYS)
        return options

    def _load_career_suggestions(self) -> dict[str, dict[str, str]]:
        """Probable Career Paths Table 2 — Suggestion on Employee Portal."""
        wb = load_workbook(SOURCE_FILES["competency"], data_only=True, read_only=True)
        if "Probable Career Paths" not in wb.sheetnames:
            return {}
        ws = wb["Probable Career Paths"]
        suggestions: dict[str, dict[str, str]] = {}
        in_table2 = False
        for cells in ws.iter_rows(values_only=True):
            first = clean(cells[0] if cells else "")
            second = clean(cells[1] if cells and len(cells) > 1 else "")
            if first == "Table 2" or "Suggestion on Employee Portal" in first:
                in_table2 = True
                continue
            if not in_table2:
                continue
            if not second or second.lower() == "role & grades":
                continue
            if not (second.startswith("BDM") or second.startswith("KAM") or second.startswith("ZM") or second.startswith("RD")):
                continue
            # cols: Role & Grades, C%XZA, KAM, ZM, RD
            suggestions[second] = {
                "kam": self._normalize_suggestion_flag(cells[3] if len(cells) > 3 else ""),
                "zm": self._normalize_suggestion_flag(cells[4] if len(cells) > 4 else ""),
            }
        return suggestions

    @staticmethod
    def _normalize_suggestion_flag(value: Any) -> str:
        text = clean(value).lower()
        if text == "yes":
            return "yes"
        if "not to be suggested" in text:
            return "hide"
        if "grey" in text or "locked" in text:
            return "grey"
        return "hide"

    def _load_competencies(self) -> list[dict[str, Any]]:
        wb = load_workbook(SOURCE_FILES["competency"], data_only=True, read_only=True)
        sheet_name = "Role-Competency Mapping" if "Role-Competency Mapping" in wb.sheetnames else "Competency"
        ws = wb[sheet_name]
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows: list[dict[str, Any]] = []
        start_row = 2 if sheet_name == "Role-Competency Mapping" else 3
        for cells in ws.iter_rows(min_row=start_row, values_only=True):
            if not cells[0] or not cells[1]:
                continue
            ideals = {}
            ideal_start = 3 if sheet_name == "Role-Competency Mapping" else 4
            for idx, header in enumerate(headers[ideal_start:], start=ideal_start):
                if header and cells[idx]:
                    value = str(cells[idx]).strip()
                    if value.upper() != "NA":
                        ideals[self._normalize_role_key(header)] = value
            rows.append(
                {
                    "skill": str(cells[0]).strip(),
                    "tag": str(cells[1]).strip(),
                    "definition": str(cells[2] or "").strip(),
                    "sales_note": str(cells[2] or "").strip(),
                    "product_note": "" if sheet_name == "Role-Competency Mapping" else str(cells[3] or "").strip(),
                    "ideals": ideals,
                }
            )
        return rows

    @staticmethod
    def _normalize_role_key(value: str) -> str:
        text = str(value).strip().replace(" ", "")
        aliases = {
            "BDM(RL2-3)": "BDM (RL2-3)",
            "BDM(RL4)": "BDM (RL4)",
            "KAM(RL2-3)": "KAM (RL2-3)",
            "KAM(RL4)": "KAM (RL4)",
            "ZM(RL4-5)": "ZM (RL4-5)",
            "ZM(RL6)": "ZM (RL6)",
            "RD(RL7-8)": "RD (RL7-8)",
        }
        return aliases.get(text, str(value).strip())

    def _load_level_definitions(self) -> dict[str, dict[str, str]]:
        wb = load_workbook(SOURCE_FILES["competency"], data_only=True, read_only=True)
        if "Competency vs Level Definitions" not in wb.sheetnames:
            return {}
        ws = wb["Competency vs Level Definitions"]
        headers = [str(value or "").strip() for value in next(ws.iter_rows(values_only=True))]
        definitions: dict[str, dict[str, str]] = {}
        for values in ws.iter_rows(min_row=2, values_only=True):
            competency = clean(values[0])
            if not competency or any(
                competency.lower().startswith(f"{level.lower()}:")
                for level in ("Beginner", "Intermediate", "Proficient", "Advanced")
            ):
                continue
            definitions[competency] = {
                headers[index]: clean(values[index])
                for index in range(1, min(len(headers), len(values)))
                if headers[index] and values[index]
            }
        return definitions

    def _load_roleplay_links(self) -> dict[str, str]:
        wb = load_workbook(SOURCE_FILES["competency"], data_only=True, read_only=True)
        links: dict[str, str] = {}
        sheet_name = "Role Plays" if "Role Plays" in wb.sheetnames else "Sheet4" if "Sheet4" in wb.sheetnames else ""
        if sheet_name:
            for skill, link in wb[sheet_name].iter_rows(min_row=2, values_only=True):
                if skill and link:
                    url = str(link).strip()
                    if url.startswith("https://") and "…" not in url and "..." not in url:
                        links[str(skill).strip()] = url
        # Fill missing competencies by repeating known links so every skill has Open Assessment.
        known = [links[row["skill"]] for row in self.competencies if links.get(row["skill"])]
        if known:
            fill_index = 0
            for row in self.competencies:
                skill = row["skill"]
                if skill not in links:
                    links[skill] = known[fill_index % len(known)]
                    fill_index += 1
        return links

    def _load_employees(self) -> dict[str, dict[str, Any]]:
        wb = load_workbook(SOURCE_FILES["darwin"], data_only=True, read_only=True)
        ws = wb.active
        headers = [str(v or "").strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows: dict[str, dict[str, Any]] = {}
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            code = clean(row.get("EMP Code"))
            if not code:
                continue
            past_years = int(row.get("Past Exp In Years") or 0)
            past_months = int(row.get("Past Exp In Months") or 0)
            rows[code] = {
                "code": code,
                "name": clean(row.get("EMP Full Name")),
                "designation": clean(row.get("Designation")),
                "level": clean(row.get("Level")),
                "location": clean(row.get("Location")),
                "state": clean(row.get("Office State")),
                "department": clean(row.get("Integration Department")),
                "function": clean(row.get("Function")),
                "manager": clean(row.get("Immediate Supervisor")),
                "manager_code": clean(row.get("Immediate Supervisor Code")),
                "rd": clean(row.get("Skip Manager")).split("(", 1)[0].strip(),
                "rd_code": clean(row.get("Skip Manager ID")),
                "role": clean(row.get("Role")),
                "past_exp_years": past_years,
                "past_exp_months": past_months,
                "total_exp_years": round(past_years + past_months / 12, 1),
                "cohort": clean(row.get("Cohort")),
            }
        return rows

    def _load_tna(self) -> dict[str, list[dict[str, str]]]:
        wb = load_workbook(SOURCE_FILES["tna"], data_only=True, read_only=True)
        ws = wb.active
        headers = [str(v or "").strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        grouped: dict[str, list[dict[str, str]]] = {}
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {key: clean(value) for key, value in zip(headers, values)}
            code = row.get("EMP Code")
            if code:
                grouped.setdefault(code, []).append(row)
        return grouped

    def _load_appraisal(self) -> dict[str, dict[str, str]]:
        wb = load_workbook(SOURCE_FILES["appraisal"], data_only=True, read_only=True)
        ws = wb.active
        headers = [str(v or "").strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows: dict[str, dict[str, str]] = {}
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {key: clean(value) for key, value in zip(headers, values)}
            code = row.get("EMP Code")
            if code:
                rows[code] = row
        return rows

    def _load_interview(self) -> dict[str, dict[str, str]]:
        path = SOURCE_FILES["interview"]
        if not path.is_file():
            return {}
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(v or "").strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows: dict[str, dict[str, str]] = {}
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {key: clean(value) for key, value in zip(headers, values)}
            code = row.get("EMP Code")
            if code:
                rows[code] = row
        return rows

    def _load_amber(self) -> dict[str, list[dict[str, str]]]:
        path = SOURCE_FILES["amber"]
        if not path.is_file():
            return {}
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(v or "").strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        grouped: dict[str, list[dict[str, str]]] = {}
        keep = {"Question", "Answer", "Mood", "Engagement Score", "Driver(Element Name)", "Driver Element Score", "Follow-up Comments"}
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = {key: clean(value) for key, value in zip(headers, values)}
            code = row.get("Emp Code")
            if code:
                grouped.setdefault(code, []).append({key: row.get(key, "") for key in keep})
        return grouped

    def _load_courses(self) -> list[dict[str, Any]]:
        path = SOURCE_FILES["courses"]
        if not path.is_file():
            return []
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        headers = [str(v or "").strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        courses: list[dict[str, Any]] = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            course_id = clean(row.get("Course ID"))
            title = _repair_catalog_text(row.get("Course Title"))
            if not course_id or not title or clean(row.get("Status")).lower() != "active":
                continue
            release = row.get("Release Date")
            courses.append(
                {
                    "id": course_id,
                    "title": title,
                    "author": _repair_catalog_text(row.get("Author")),
                    "release_date": release.strftime("%Y-%m-%d") if hasattr(release, "strftime") else clean(release),
                    "level": clean(row.get("Level")).lower(),
                    "duration": clean(row.get("Duration")),
                    "category": clean(row.get("Category")),
                    "subjects": _repair_catalog_text(row.get("Subjects")),
                    "description": _repair_catalog_text(row.get("Description")),
                    "url": clean(row.get("SSO URL")) or clean(row.get("Course URL")),
                    "thumbnail": clean(row.get("Large Thumbnail")),
                }
            )
        return courses


def _repair_catalog_text(value: Any) -> str:
    text = clean(value)
    if not any(marker in text for marker in ("â", "Ã", "Â")):
        return text
    for encoding in ("cp1252", "latin-1"):
        try:
            return text.encode(encoding).decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
    replacements = {
        "â€™": "’", "â€˜": "‘", "â€œ": "“", "â€\x9d": "”",
        "â€”": "—", "â€“": "–", "â€¦": "…", "Â": "",
    }
    for broken, repaired in replacements.items():
        text = text.replace(broken, repaired)
    return text
