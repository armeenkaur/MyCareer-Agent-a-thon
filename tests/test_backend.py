from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from skillsync_ai.backend import BackendError, MyCareerBackend
from skillsync_ai.data_sources import WorkbookData
from skillsync_ai.database import Database, generated_password, utc_now


class BackendWorkflowTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.data = WorkbookData()

    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.db = Database(f"{self.temp.name}/test.db")
        self.backend = MyCareerBackend(self.data, self.db)
        self.admin = self.db.authenticate("ADMIN", "admin", "Admin")
        assert self.admin is not None

    def tearDown(self) -> None:
        self.temp.cleanup()

    def test_role_scoped_accounts_and_generated_passwords(self) -> None:
        employee = self.data.employees["MMT1001"]
        account = self.db.authenticate("MMT1001", "employee", generated_password(employee["name"]))
        self.assertIsNotNone(account)
        self.assertIsNone(self.db.authenticate("MMT1001", "employee", "wrong"))

        # Dinesh keeps separate ZM and RD account rows under the same login ID.
        self.assertIsNotNone(self.db.authenticate("MMT11043", "zm", "Dinesh"))
        self.assertIsNotNone(self.db.authenticate("MMT11043", "rd", "Dinesh"))

    def test_runtime_cache_clear_preserves_source_and_workflow_data(self) -> None:
        self.db.create_session(int(self.admin["id"]))
        with self.db.transaction() as connection:
            connection.execute(
                "INSERT INTO curated_evidence(employee_code,competency,evidence_json,generated_at) VALUES(?,?,?,?)",
                ("MMT1001", "Communication", "{}", utc_now()),
            )
            connection.execute(
                """
                INSERT INTO course_recommendations(
                    employee_code,target_key,competency,current_level,target_level,
                    candidate_ids_json,courses_json,generated_at
                ) VALUES(?,?,?,?,?,?,?,?)
                """,
                ("MMT1001", "test", "Communication", "Beginner", "Intermediate", "[]", "[]", utc_now()),
            )

        self.db.clear_runtime_cache()

        with self.db.connect() as connection:
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM sessions").fetchone()[0], 0)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM curated_evidence").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM course_recommendations").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM employees").fetchone()[0], 44)

    def test_phase_gate_blocks_login_until_admin_opens_phase(self) -> None:
        employee = self.data.employees["MMT1001"]
        with self.assertRaises(BackendError) as error:
            self.backend.login(employee["manager_code"], generated_password(employee["manager"]), role="zm")
        self.assertEqual(error.exception.code, "phase_closed")

        self.backend.open_phase(self.admin, "zm")
        result = self.backend.login(employee["manager_code"], generated_password(employee["manager"]), role="zm")
        self.assertEqual(result["user"]["role"], "zm")

    def test_roleless_login_defaults_dual_zm_rd_to_zm(self) -> None:
        self.backend.open_phase(self.admin, "zm")
        result = self.backend.login("MMT11043", "Dinesh")
        self.assertEqual(result["user"]["role"], "zm")
        self.assertEqual(result["user"]["available_roles"], ["zm", "rd"])

    def test_switch_role_and_password_sync_across_dual_accounts(self) -> None:
        self.backend.open_phase(self.admin, "zm")
        self.backend.open_phase(self.admin, "rd", override=True)
        # Opening RD closes ZM phase; re-open ZM for dual switch test.
        with self.db.transaction() as connection:
            connection.execute(
                "UPDATE phases SET status='open', closed_at=NULL WHERE phase='zm'"
            )
        zm_login = self.backend.login("MMT11043", "Dinesh")
        self.assertEqual(zm_login["user"]["role"], "zm")
        switched = self.backend.switch_role(zm_login["user"], zm_login["token"], "rd")
        self.assertEqual(switched["user"]["role"], "rd")

        rd_user = self.db.authenticate("MMT11043", "rd", "Dinesh")
        assert rd_user is not None
        self.backend.change_password(rd_user, "Dinesh", "SyncedPass1")
        self.assertIsNotNone(self.db.authenticate("MMT11043", "zm", "SyncedPass1"))
        self.assertIsNotNone(self.db.authenticate("MMT11043", "rd", "SyncedPass1"))
        self.assertIsNone(self.db.authenticate("MMT11043", "zm", "Dinesh"))

    def test_switch_role_shows_phase_closed_when_target_not_open(self) -> None:
        self.backend.open_phase(self.admin, "zm")
        zm_login = self.backend.login("MMT11043", "Dinesh")
        with self.assertRaises(BackendError) as error:
            self.backend.switch_role(zm_login["user"], zm_login["token"], "rd")
        self.assertEqual(error.exception.code, "phase_closed")
        self.assertEqual(self.db.session_user(zm_login["token"])["role"], "zm")

    def test_badge_catalog_two_hour_and_five_day_streak(self) -> None:
        from datetime import datetime, timedelta, timezone
        from skillsync_ai.backend import BADGE_CATALOG

        ids = {badge["id"] for badge in BADGE_CATALOG}
        self.assertEqual(
            ids,
            {"two_hour_club", "ten_hour_club", "five_day_streak", "full_circuit", "gap_closer"},
        )
        self.assertNotIn("hours_stacked", ids)
        self.assertNotIn("cohort_crown", ids)

        code = "MMT1001"
        today = datetime.now(timezone.utc).date()
        for offset in range(5):
            day = (today - timedelta(days=offset)).isoformat()
            self.backend.record_learning_day(code, day)
        self.assertGreaterEqual(self.backend.learning_streak(code), 5)

        badges = self.backend._sync_badges_for_row(
            {
                "employee_code": code,
                "learning_hours": 2.5,
                "completions": 0,
                "focus_areas": 1,
                "severity_band": 3,
                "rank": 2,
                "full_circuit": False,
            }
        )
        earned = {badge["id"] for badge in badges}
        self.assertIn("two_hour_club", earned)
        self.assertIn("five_day_streak", earned)
        self.assertNotIn("first_mile", earned)
        self.assertNotIn("cohort_crown", earned)

    def test_phase_override_does_not_report_incomplete_previous_phase_as_complete(self) -> None:
        self.backend.open_phase(self.admin, "rd", override=True)

        self.assertEqual(self.backend.phase("zm")["status"], "closed")
        self.assertEqual(self.backend.phase("rd")["status"], "open")
        self.assertFalse(self.backend.phase("zm")["progress"]["is_complete"])

    def test_rd_cannot_rate_employee_before_zm_submission(self) -> None:
        employee = self.data.employees["MMT1001"]
        self.backend.open_phase(self.admin, "rd", override=True)
        rd = self.db.authenticate(employee["rd_code"], "rd", generated_password(employee["rd"]))
        assert rd is not None

        with self.assertRaises(BackendError) as error:
            self.backend.save_assessment(
                rd,
                employee["code"],
                {competency: "Intermediate" for competency in self.backend.competencies},
                submit=False,
            )

        self.assertEqual(error.exception.code, "zm_assessment_required")

    def test_rd_validation_does_not_run_evidence_agent_before_zm_submission(self) -> None:
        employee = self.data.employees["MMT1001"]
        rd = self.db.authenticate(employee["rd_code"], "rd", generated_password(employee["rd"]))
        assert rd is not None

        with self.assertRaises(BackendError) as error:
            self.backend.rd_validation_context(rd, employee["code"])

        self.assertEqual(error.exception.code, "zm_assessment_required")
        self.assertEqual(self.backend.agent_audit(), [])

    @patch("skillsync_ai.backend.learning._rank_all_with_agent", return_value=({}, "test ranker"))
    def test_zm_then_rd_submission_makes_rd_profile_final(self, ranker) -> None:
        employee = self.data.employees["MMT1001"]
        self.backend.open_phase(self.admin, "zm")
        zm = self.db.authenticate(employee["manager_code"], "zm", generated_password(employee["manager"]))
        assert zm is not None
        zm_ratings = {competency: "Beginner" for competency in self.backend.competencies}
        result = self.backend.save_assessment(zm, "MMT1001", zm_ratings, submit=True, career_recommendation="continue")
        self.assertEqual(result["status"], "submitted")

        self.backend.open_phase(self.admin, "rd", override=True)
        rd = self.db.authenticate(employee["rd_code"], "rd", generated_password(employee["rd"]))
        assert rd is not None
        rd_ratings = {competency: "Intermediate" for competency in self.backend.competencies}
        result = self.backend.save_assessment(rd, "MMT1001", rd_ratings, submit=True, career_recommendation="continue")
        self.assertEqual(result["status"], "submitted")
        self.assertEqual(self.backend.final_profile("MMT1001"), rd_ratings)
        expected_agent_runs = sum(
            bool(target["gaps"]) for target in self.backend.recommendation_targets("MMT1001")
        )
        self.assertEqual(ranker.call_count, expected_agent_runs)
        with self.db.transaction() as connection:
            for kind in ("functional", "behavioural"):
                connection.execute(
                    """
                    INSERT INTO voice_roleplay_sessions(
                        employee_code,kind,status,scores_json,error,updated_at
                    ) VALUES(?,?, 'completed','{}','',?)
                    """,
                    ("MMT1001", kind, utc_now()),
                )
            connection.execute(
                """
                INSERT INTO career_choices(employee_code,aspiration_role,target_key,locked_at)
                VALUES(?,?,?,?)
                """,
                ("MMT1001", "kam", "KAM (RL4)", utc_now()),
            )
        self.assertTrue(self.backend.recommendations("MMT1001")["ready"])

        with self.db.transaction() as connection:
            connection.execute("DELETE FROM voice_roleplay_sessions WHERE employee_code=?", ("MMT1001",))
            connection.execute("DELETE FROM career_choices WHERE employee_code=?", ("MMT1001",))
        with self.assertRaises(BackendError) as blocked:
            self.backend.recommendations("MMT1001")
        self.assertEqual(blocked.exception.code, "lattice_locked")

    @patch("skillsync_ai.backend.learning._rank_all_with_agent", return_value=({}, "test ranker"))
    def test_admin_can_reset_zm_and_rd_assessments(self, _ranker) -> None:
        employee = self.data.employees["MMT1001"]
        self.backend.open_phase(self.admin, "zm")
        zm = self.db.authenticate(employee["manager_code"], "zm", generated_password(employee["manager"]))
        assert zm is not None
        ratings = {competency: "Intermediate" for competency in self.backend.competencies}
        self.backend.save_assessment(zm, "MMT1001", ratings, submit=True, career_recommendation="continue")
        self.backend.open_phase(self.admin, "rd", override=True)
        rd = self.db.authenticate(employee["rd_code"], "rd", generated_password(employee["rd"]))
        assert rd is not None
        self.backend.save_assessment(rd, "MMT1001", ratings, submit=True, career_recommendation="continue")
        self.assertIsNotNone(self.backend.final_profile("MMT1001"))

        rd_only = self.backend.reset_manager_assessments(self.admin, "MMT1001", scope="rd")
        self.assertEqual(rd_only["cleared"], ["rd"])
        self.assertEqual(rd_only["zm_status"], "submitted")
        self.assertEqual(rd_only["rd_status"], "not_started")
        self.assertIsNone(self.backend.final_profile("MMT1001"))
        self.assertEqual(self.backend.assessment("MMT1001", "zm")["status"], "submitted")

        both = self.backend.reset_manager_assessments(self.admin, "MMT1001", scope="zm")
        self.assertIn("zm", both["cleared"])
        self.assertEqual(both["zm_status"], "not_started")
        self.assertEqual(both["rd_status"], "not_started")
        self.assertIsNone(self.backend.assessment("MMT1001", "zm"))

    @patch("skillsync_ai.backend.learning._rank_all_with_agent", return_value=({}, "test ranker"))
    def test_confidence_is_deterministic_and_uses_zm_plus_ai(self, _ranker) -> None:
        employee = self.data.employees["MMT1001"]
        self.backend.open_phase(self.admin, "zm")
        zm = self.db.authenticate(employee["manager_code"], "zm", generated_password(employee["manager"]))
        assert zm is not None
        ratings = {competency: "Proficient" for competency in self.backend.competencies}
        self.backend.save_assessment(zm, "MMT1001", ratings, submit=True, career_recommendation="continue")
        self.backend.open_phase(self.admin, "rd", override=True)
        rd = self.db.authenticate(employee["rd_code"], "rd", generated_password(employee["rd"]))
        assert rd is not None
        self.backend.save_assessment(rd, "MMT1001", ratings, submit=True, career_recommendation="continue")
        with self.db.transaction() as connection:
            for competency in self.backend.competencies:
                connection.execute(
                    """
                    INSERT INTO roleplay_assessments(
                        employee_code,competency,status,ai_proficiency,updated_at
                    ) VALUES(?,?, 'completed','Proficient',?)
                    """,
                    ("MMT1001", competency, utc_now()),
                )
        confidence = self.backend.confidence("MMT1001")
        self.assertEqual(confidence["status"], "complete")
        self.assertEqual(confidence["score"], 100.0)

    def test_roleplay_assessment_evidence_is_admin_only(self) -> None:
        screenshot = Path(self.temp.name) / "communication.png"
        screenshot.write_bytes(b"private screenshot")
        with self.db.transaction() as connection:
            connection.execute(
                """
                INSERT INTO roleplay_assessments(
                    employee_code,competency,filename,file_path,status,ai_proficiency,
                    rationale,ocr_text,error,updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    "MMT1001",
                    "Communication",
                    screenshot.name,
                    str(screenshot),
                    "completed",
                    "Proficient",
                    "Observed behavior",
                    "Private OCR transcript",
                    "",
                    utc_now(),
                ),
            )

        employee_view = next(
            row for row in self.backend.roleplays("MMT1001") if row["competency"] == "Communication"
        )
        for private_field in ("ai_proficiency", "rationale", "ocr_text", "filename", "file_path"):
            self.assertNotIn(private_field, employee_view)

        admin_view = self.backend.admin_roleplays(self.admin, "MMT1001", "Communication")
        communication = next(
            row for row in admin_view["roleplays"] if row["competency"] == "Communication"
        )
        self.assertEqual(communication["ai_proficiency"], "Proficient")
        self.assertEqual(communication["rationale"], "Observed behavior")
        self.assertEqual(admin_view["screenshot"]["content_base64"], "cHJpdmF0ZSBzY3JlZW5zaG90")

        employee = self.db.authenticate("MMT1001", "employee", generated_password(self.data.employees["MMT1001"]["name"]))
        assert employee is not None
        with self.assertRaises(BackendError) as error:
            self.backend.admin_roleplays(employee, "MMT1001")
        self.assertEqual(error.exception.code, "forbidden")

    @patch("skillsync_ai.backend.learning._rank_all_with_agent")
    def test_aspiration_lock_and_backend_filtered_course_candidates(self, ranker) -> None:
        ranker.side_effect = lambda groups, *_args: (
            {key: _first_two(value["candidates"]) for key, value in groups.items()},
            "test ranker",
        )
        employee = self.data.employees["MMT1002"]
        self.backend.open_phase(self.admin, "zm")
        zm = self.db.authenticate(employee["manager_code"], "zm", generated_password(employee["manager"]))
        assert zm is not None
        ratings = {competency: "Advanced" for competency in self.backend.competencies}
        self.backend.save_assessment(zm, "MMT1002", ratings, submit=True, career_recommendation="continue")
        self.backend.open_phase(self.admin, "rd", override=True)
        rd = self.db.authenticate(employee["rd_code"], "rd", generated_password(employee["rd"]))
        assert rd is not None
        self.backend.save_assessment(rd, "MMT1002", ratings, submit=True, career_recommendation="continue")
        self.backend.open_phase(self.admin, "employee", override=True)
        with self.db.transaction() as connection:
            for kind in ("functional", "behavioural"):
                connection.execute(
                    """
                    INSERT INTO voice_roleplay_sessions(
                        employee_code,kind,status,scores_json,error,updated_at
                    ) VALUES(?,?, 'completed','{}','',?)
                    """,
                    ("MMT1002", kind, utc_now()),
                )
            for competency in self.backend.competencies:
                connection.execute(
                    """
                    INSERT INTO roleplay_assessments(
                        employee_code,competency,status,ai_proficiency,updated_at
                    ) VALUES(?,?, 'completed','Advanced',?)
                    """,
                    ("MMT1002", competency, utc_now()),
                )
        user = self.db.authenticate("MMT1002", "employee", generated_password(employee["name"]))
        assert user is not None
        state = self.backend.choose_career(user, "kam")
        self.assertEqual(state["choice"]["aspiration_role"], "kam")
        with self.assertRaises(BackendError) as error:
            self.backend.choose_career(user, "kam")
        self.assertEqual(error.exception.code, "career_locked")

        with self.db.connect() as connection:
            rows = connection.execute(
                "SELECT candidate_ids_json FROM course_recommendations WHERE employee_code='MMT1002'"
            ).fetchall()
        # Candidate rows can be empty when final profile already meets the target; no general course can enter either way.
        self.assertTrue(all("general" not in row["candidate_ids_json"] for row in rows))

    def test_leaderboard_ranks_hours_pct_then_courses_completed(self) -> None:
        now = utc_now()
        with self.db.transaction() as connection:
            for code in ("MMT1002", "MMT1004"):
                cursor = connection.execute(
                    """
                    INSERT INTO assessments(
                        employee_code,assessor_role,assessor_login_id,status,created_at,updated_at,submitted_at
                    ) VALUES(?, 'rd', 'TEST-RD', 'submitted', ?, ?, ?)
                    """,
                    (code, now, now, now),
                )
                for competency in self.backend.competencies:
                    connection.execute(
                        "INSERT INTO assessment_ratings(assessment_id,competency,proficiency) VALUES(?,?,'Beginner')",
                        (cursor.lastrowid, competency),
                    )
                for idx, course_id in enumerate((f"{code}-A", f"{code}-B")):
                    connection.execute(
                        "INSERT INTO learning_selections(employee_code,competency,course_id,selected_at) VALUES(?,?,?,?)",
                        (code, "Communication", course_id, now),
                    )
                    status = "completed" if (code == "MMT1002" or idx == 0) else "not_started"
                    pct = 100 if status == "completed" else 0
                    connection.execute(
                        """
                        INSERT INTO course_progress(employee_code,course_id,status,progress_pct,launched_at,completed_at)
                        VALUES(?,?,?,?,?,?)
                        """,
                        (code, course_id, status, pct, now, now if status == "completed" else None),
                    )
                # Same LinkedIn hours; ranking must use journey % / courses completed, not raw hours.
                connection.execute(
                    "INSERT INTO linkedin_activity(employee_code,learning_hours,completions,synced_at) VALUES(?,?,?,?)",
                    (code, 3.5, 1, now),
                )
            # Force equal % via known catalog durations by using progress-only fallback:
            # no catalog ids → hours_pct from courses_completed/courses_total.
            # MMT1002: 2/2 = 100%, MMT1004: 1/2 = 50%
        payload = self.backend.leaderboard(self.admin, force_refresh=True)
        rows = [row for row in payload["leaderboard"] if row["employee_code"] in {"MMT1002", "MMT1004"}]
        self.assertEqual(len(rows), 2)
        by_code = {row["employee_code"]: row for row in rows}
        self.assertEqual(by_code["MMT1002"]["hours_pct"], 100.0)
        self.assertEqual(by_code["MMT1004"]["hours_pct"], 50.0)
        self.assertEqual(by_code["MMT1002"]["rank"], 1)
        self.assertEqual(by_code["MMT1004"]["rank"], 2)

        # Same % → higher courses_completed ranks first
        with self.db.transaction() as connection:
            connection.execute(
                "UPDATE course_progress SET status='completed', progress_pct=100 WHERE employee_code='MMT1004' AND course_id='MMT1004-B'"
            )
            connection.execute(
                "DELETE FROM course_progress WHERE employee_code='MMT1002' AND course_id='MMT1002-B'"
            )
            connection.execute(
                "DELETE FROM learning_selections WHERE employee_code='MMT1002' AND course_id='MMT1002-B'"
            )
        payload = self.backend.leaderboard(self.admin, force_refresh=True)
        rows = [row for row in payload["leaderboard"] if row["employee_code"] in {"MMT1002", "MMT1004"}]
        by_code = {row["employee_code"]: row for row in rows}
        self.assertEqual(by_code["MMT1002"]["hours_pct"], by_code["MMT1004"]["hours_pct"])
        self.assertEqual(by_code["MMT1004"]["courses_completed"], 2)
        self.assertEqual(by_code["MMT1002"]["courses_completed"], 1)
        self.assertEqual(by_code["MMT1004"]["rank"], 1)
        self.assertEqual(by_code["MMT1002"]["rank"], 2)

        # Full tie on % + courses → shared rank
        with self.db.transaction() as connection:
            connection.execute(
                "INSERT INTO learning_selections(employee_code,competency,course_id,selected_at) VALUES(?,?,?,?)",
                ("MMT1002", "Communication", "MMT1002-B", now),
            )
            connection.execute(
                """
                INSERT INTO course_progress(employee_code,course_id,status,progress_pct,launched_at,completed_at)
                VALUES(?,?,?,?,?,?)
                """,
                ("MMT1002", "MMT1002-B", "completed", 100, now, now),
            )
        payload = self.backend.leaderboard(self.admin, force_refresh=True)
        rows = [row for row in payload["leaderboard"] if row["employee_code"] in {"MMT1002", "MMT1004"}]
        self.assertEqual({row["rank"] for row in rows}, {1})

    def test_course_frontend_contract_contains_required_catalog_fields(self) -> None:
        row = self.backend._course_contract(
            {"title": "Example", "url": "https://example.com", "author": "Provider", "duration": "01:00:00"},
            "Communication",
            "Beginner",
            "Intermediate",
        )
        self.assertEqual(row["provider"], "Provider")
        self.assertEqual(row["source_type"], "LinkedIn Learning")
        self.assertEqual(row["competency"], "Communication")
        self.assertEqual(row["supported_proficiency_movement"], {"from": "Beginner", "to": "Intermediate"})

    def test_frontend_employee_summary_and_final_profile_contract(self) -> None:
        employee = self.data.employees["MMT1001"]
        with self.db.transaction() as connection:
            cursor = connection.execute(
                """
                INSERT INTO assessments(employee_code,assessor_role,assessor_login_id,status,created_at,updated_at,submitted_at)
                VALUES('MMT1001','rd','TEST-RD','submitted',?,?,?)
                """,
                (utc_now(), utc_now(), utc_now()),
            )
            for competency in self.backend.competencies:
                connection.execute(
                    "INSERT INTO assessment_ratings(assessment_id,competency,proficiency) VALUES(?,?,'Intermediate')",
                    (cursor.lastrowid, competency),
                )
        zm = self.db.authenticate(employee["manager_code"], "zm", generated_password(employee["manager"]))
        assert zm is not None
        summary = next(row for row in self.backend.employee_summaries(zm) if row["employee_code"] == "MMT1001")
        self.assertTrue(summary["final_profile_available"])
        profile = self.backend.profile_for_user(zm, "MMT1001")
        self.assertEqual(profile["status"], "final")
        self.assertEqual(len(profile["ratings"]), 7)

    def test_agent_audit_contract_lists_recorded_agent(self) -> None:
        self.backend._audit("MMT1001", "Evidence Curator Agent", "Communication", "TNA", {"evidence": []}, "ok")
        rows = self.backend.agent_audit()
        self.assertEqual(rows[0]["agent"], "Evidence Curator Agent")
        self.assertEqual(rows[0]["output"], {"evidence": []})

    def test_assessment_template_and_upload_for_zm(self) -> None:
        from io import BytesIO

        from openpyxl import Workbook, load_workbook

        employee = self.data.employees["MMT1001"]
        self.backend.open_phase(self.admin, "zm")
        zm = self.backend.login(employee["manager_code"], generated_password(employee["manager"]), role="zm")["user"]

        payload, filename = self.backend.download_assessment_template(zm)
        self.assertTrue(filename.endswith(".xlsx"))
        wb = load_workbook(BytesIO(payload), data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers[0], "Employee Code")
        self.assertEqual(headers[1], "Employee Name")
        for competency in self.backend.competencies:
            self.assertIn(competency, headers)
        codes = {
            str(row[0]).strip()
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row and row[0] and str(row[0]).strip().upper().startswith("MMT")
        }
        self.assertIn("MMT1001", codes)

        # Build upload with all seven ratings → submit
        upload = Workbook()
        sheet = upload.active
        sheet.append(["Employee Code", "Employee Name", *self.backend.competencies])
        sheet.append(["MMT1001", employee["name"], *(["Intermediate"] * len(self.backend.competencies))])
        buffer = BytesIO()
        upload.save(buffer)
        result = self.backend.upload_assessment_workbook(zm, "filled.xlsx", buffer.getvalue())
        self.assertEqual(result["summary"]["applied"], 1)
        self.assertEqual(result["summary"]["errors"], 0)
        assessment = self.backend.assessment("MMT1001", "zm")
        self.assertEqual(assessment["status"], "submitted")
        self.assertEqual(len(assessment["ratings"]), 7)

        # Locked row skipped on re-upload
        again = self.backend.upload_assessment_workbook(zm, "filled.xlsx", buffer.getvalue())
        self.assertEqual(again["summary"]["applied"], 0)
        self.assertGreaterEqual(again["summary"]["skipped"], 1)

        # Out-of-scope code errors
        bad = Workbook()
        bad_sheet = bad.active
        bad_sheet.append(["Employee Code", "Employee Name", *self.backend.competencies])
        bad_sheet.append(["MMT99999", "Nobody", *(["Beginner"] * len(self.backend.competencies))])
        bad_buf = BytesIO()
        bad.save(bad_buf)
        bad_result = self.backend.upload_assessment_workbook(zm, "bad.xlsx", bad_buf.getvalue())
        self.assertGreaterEqual(bad_result["summary"]["errors"], 1)

    def test_assessment_upload_partial_saves_draft_and_rd_requires_zm(self) -> None:
        from io import BytesIO

        from openpyxl import Workbook

        employee = self.data.employees["MMT1001"]
        self.backend.open_phase(self.admin, "zm")
        zm = self.db.authenticate(employee["manager_code"], "zm", generated_password(employee["manager"]))
        assert zm is not None

        partial = Workbook()
        sheet = partial.active
        sheet.append(["Employee Code", "Employee Name", *self.backend.competencies])
        values = [""] * len(self.backend.competencies)
        values[0] = "Beginner"
        values[1] = "Intermediate"
        sheet.append(["MMT1001", employee["name"], *values])
        buf = BytesIO()
        partial.save(buf)
        result = self.backend.upload_assessment_workbook(zm, "partial.xlsx", buf.getvalue())
        self.assertEqual(result["summary"]["applied"], 1)
        self.assertEqual(result["applied"][0]["submitted"], False)
        self.assertEqual(self.backend.assessment("MMT1001", "zm")["status"], "draft")

        # Opening RD closes ZM; RD upload still blocked because ZM is draft, not submitted.
        self.backend.open_phase(self.admin, "rd", override=True)
        rd = self.db.authenticate(employee["rd_code"], "rd", generated_password(employee["rd"]))
        assert rd is not None
        rd_book = Workbook()
        rd_sheet = rd_book.active
        rd_sheet.append(["Employee Code", "Employee Name", *self.backend.competencies])
        rd_sheet.append(["MMT1001", employee["name"], *(["Proficient"] * len(self.backend.competencies))])
        rd_buf = BytesIO()
        rd_book.save(rd_buf)
        rd_result = self.backend.upload_assessment_workbook(rd, "rd.xlsx", rd_buf.getvalue())
        self.assertGreaterEqual(rd_result["summary"]["errors"], 1)

    def test_voice_roleplay_sessions_unlock_lattice(self) -> None:
        employee = self.data.employees["MMT1001"]
        self.backend.open_phase(self.admin, "employee", override=True)
        user = self.db.authenticate("MMT1001", "employee", generated_password(employee["name"]))
        assert user is not None
        self.assertFalse(self.backend.lattice_unlocked("MMT1001"))

        start = self.backend.start_voice_roleplay(user, "functional")
        self.assertEqual(start["kind"], "functional")
        self.assertIn("session_id", start)
        ticket = self.backend.voice_roleplay_ticket(start["session_id"], user)
        self.assertEqual(ticket["kind"], "functional")

        functional_ratings = {
            "Consultative Selling": {"level": "Intermediate", "confidence": 0.8},
            "Data Analytics": {"level": "Beginner", "confidence": 0.7},
            "Stakeholder Relationship": {"level": "Proficient", "confidence": 0.9},
            "Communication": {"level": "Intermediate", "confidence": 0.6},
            "Executive Presence": {"level": "Proficient", "confidence": 0.5},
            "Ownership & Accountability": None,
            "Team Management": None,
        }
        done = self.backend.complete_voice_roleplay(start["session_id"], "MMT1001", functional_ratings)
        self.assertEqual(done["status"], "completed")
        self.assertFalse(done["lattice_unlocked"])

        start_b = self.backend.start_voice_roleplay(user, "behavioural")
        behavioural_ratings = {
            "Communication": {"level": "Proficient", "confidence": 0.9},
            "Ownership & Accountability": {"level": "Intermediate", "confidence": 0.8},
            "Team Management": {"level": "Beginner", "confidence": 0.7},
            "Executive Presence": {"level": "Advanced", "confidence": 0.85},
            "Stakeholder Relationship": {"level": "Advanced", "confidence": 0.8},
            "Data Analytics": None,
            "Consultative Selling": None,
        }
        done_b = self.backend.complete_voice_roleplay(start_b["session_id"], "MMT1001", behavioural_ratings)
        self.assertTrue(done_b["lattice_unlocked"])
        self.assertTrue(self.backend.lattice_unlocked("MMT1001"))
        sessions = self.backend.voice_roleplay_sessions("MMT1001")
        self.assertEqual({row["status"] for row in sessions}, {"completed"})
        roleplays = {row["competency"]: row for row in self.backend.roleplays("MMT1001", include_private=True)}
        self.assertEqual(roleplays["Data Analytics"]["ai_proficiency"], "Beginner")
        self.assertEqual(roleplays["Executive Presence"]["ai_proficiency"], "Advanced")
        self.assertEqual(roleplays["Ownership & Accountability"]["ai_proficiency"], "Intermediate")
        # Merged Communication: Intermediate@0.6 + Proficient@0.9 → Proficient
        self.assertEqual(roleplays["Communication"]["ai_proficiency"], "Proficient")
        self.assertIn("confidence-weighted merge", roleplays["Communication"]["rationale"])

def _first_two(candidates):
    return [dict(row) for row in candidates[:2]]


if __name__ == "__main__":
    unittest.main()
